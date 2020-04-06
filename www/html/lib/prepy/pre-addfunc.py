'''#########################################################################
# File Name: datafunction.py
# Author: Tang
# mail: tanzhengtang@163.com
# Created Time: 2019年09月13日 星期五 14时39分50秒
# Last write Time:2019年10月14日 星期一 20点50分
# New Problem 无法做到对每个Excel进行读入，同时只支持Excel格式，函数式编程,数据库中有重复表存在时会报错。
# 待解决：解决数据表主键冲突问题 ，目前暂时设置为无主键 
# 待解决：别名表、PMID表、数据表
# 待解决：环境改变需要从写的地方太多了
#########################################################################'''
import openpyxl
import sys
import os
import re
import MySQLdb as db
# 报错机制未解决
'''删除表格中pmid，title ,orig_text的空值，注意替换表格值'''


def delnon():
    wb = openpyxl.load_workbook(
        '/home/tang/sju3734/project1/www/html/lib/brca.xlsx')
    ws = wb.get_sheet_by_name('brca_v1')
    print(ws.max_row)
    for i in range(1, ws.max_row):
        if ws["G%d" % i].value == None and ws["H%d" % i].value == None:
            ws["G%d" % i].value = ws["G%d" % (i-1)].value
            ws["H%d" % i].value = ws["H%d" % (i-1)].value
            print("G%d.value loaded\n" % i)
    wb.save('/home/tang/sju3734/project1/www/html/lib/brca.xlsx')
# 创建函数


def cretab(tab_row):
    f = 0
    for keyname in tab_row:
        # print("tabname is %s\n"%tabname)
        key_val = str(keyname.value).replace(" ", "")
        # print("now is %s\n"%key_val)
        cresql = ""
        if f > 0 and f < 6:
            cresql = """alter table %s add %s varchar(30)""" % (
                tabname, key_val)
            f = f+1
        elif f >= 6:
            cresql = """alter table %s add %s text""" % (tabname, key_val)
        else:
            cresql = """CREATE TABLE %s(
            %s CHAR(10))""" % (tabname, key_val)  # 改写：去掉主键约束，时间2019.9.17 14点32分
            f = f+1
            print("create the cols...")
        # print("%s\n"%cresql)
        try:
            cur.execute(cresql)
            print("...")
        except:
            con.rollback()
            print("wrong!plesea check the code\n")
    if f == 6:
        print("successfully finished!\n")

# 增加函数


def intab(tab_row):
    row_value = []
    for i in tab_row:
        if i.value == None:  # 如果cell值为None，将存入Null字符，注意以后进行查询时，空格字符的匹配
            row_value.append("'Null'")
        else:
            row_value.append("'%s'" % str(
                i.value).replace("'", "\\'"))  # 替换字符中的单引号
    row_value = ",".join(row_value)
    insql = """INSERT INTO lcbb3734.%s
         VALUES (%s);""" % (tabname, row_value)
    print("%s\n" % insql)
    cur.execute(insql)
    return 1


'''def intab_alias(tab_rwo): #插入基因别名的函数，还未修改s
    row_value=[]
    for i in tab_row:  
        if i.value==None:   
            row_value.append("'Null'")
        else :
            row_value.append("'%s'"%str(i.value).replace("'","\\'"))#替换字符中的单引号
    row_value=",".join(row_value)   
    insql= """INSERT INTO Brca.%s
         VALUES (%s);"""%(tabname,row_value)
    print("%s\n"%insql)     
    cur.execute(insql)
    return 1 '''


# host,usr,password,database 为以后改写提供参数接口
host = "localhost"
usr = "lcbb3734"
password = "3734"
database = "lcbb3734"
con = db.connect(host, usr, password, database,
                 use_unicode=True, charset="utf8")  # 切记检查编码以及添加编码。
cur = con.cursor()
wb = openpyxl.load_workbook(
    '/home/tang/sju3734/project1/www/html/lib/brca.xlsx')  # 以后考虑改写时自行查找当前路径的Excel表。
tabname = ""  # 该变量为全局变量，在cretab函数以及intab函数中均有用到，请注意在函数内部中不要出现修改该变量的执行程式
f = 0  # 循环下标
for sheet in wb:
    tabname = wb.sheetnames[f]
    f = f+1  # 此时f加1，为取出下一张excel的数据表做准备  sheet为实际的数据薄参数
    # print(tabname)
    rows = sheet.max_row
    print("max row is %d\n" % rows)
    tab_row = (x for x in sheet)
    # 以下注释代码为原始版本，易于看清执行过程，执行效率相对低，用于修改函数
    # for x in range(1,rows):
    #     if x!=1:
    #       if intab(next(tab_row))==1:
    #           print("now is inserting row of %d\n"%x)
    #       else :
    #           print("wrong!check the code!\n")
    #           break
    #     else:
    #       print("now is cretabing row of %d\n"%x)
    #       cretab(next(tab_row))

    # #该版本为重写后，效率相对高
    cretab(next(tab_row))
    [intab(row) for row in tab_row]

con.commit()
con.close()
