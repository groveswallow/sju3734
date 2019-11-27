'''
@Author: Tang
@Date: 2019-11-25 20:26:16
@LastEditors: Tang
@LastEditTime: 2019-11-26 21:14:21
@Description: 填充空格，以及删除表格头两列,填充第一行，删除多余的空白行
未解决：１．遇到多余的空行停止插入。
'''
import openpyxl,sys,os,re
def delspace(filedir):
    wb=openpyxl.load_workbook(filedir)
    ws=wb.get_sheet_by_name(wb.sheetnames[0])
    ws.delete_cols(1,2) #删除从第一列开始算的2列内容
    ws.insert_rows(0) #插入一行，把表头补全
    row=['A','B','C','D','E','F','G','H']
    head=['pmid','tf','characteristics','gene','regulation_type','cancer _hallmark','original _text','title']
    for i in range(9):
        ws["%s%d"%(row[i],i)].value=head[i] 
    print(ws["%s%d"%('A',2)].value)
    for i in range(1,ws.max_row):
        flag=0
        for x in row:
            if  ws["%s%d"%(x,i)].value==None:
                # if ws["%s%d"%(x,i)] is tuple:
                #     print(ws["%s%d"%(x,i)])
                #     print("\n")
                #     print("%s%d"%(x,i))
                ws["%s%d"%(x,i)].value=ws["%s%d"%(x,i-1)].value
                flag=flag+1
        print("the row of %d is over.\n"%i)
    wb.save(filedir)

if __name__=="__main__":
    path=sys.argv[1]
    file_dir=os.listdir(path)
    for i in file_dir:
        filedir=os.path.join(path,i)
        delspace(filedir)
