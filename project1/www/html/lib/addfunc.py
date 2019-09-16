'''#########################################################################
# File Name: datafunction.py
# Author: Tang
# mail: tanzhengtang@163.com
# Created Time: 2019年09月13日 星期五 14时39分50秒
# Last write Time:2019年9月16日 星期一 20点56分
# New Problem 无法做到对每个Excel进行读入，同时只支持Excel格式，函数式编程,数据库中有重复表存在时会报错。
#########################################################################'''
import openpyxl,sys,os,re
import MySQLdb as db
#报错机制未解决 
#创建函数
def cretab(tab_row):
    f=0
    for keyname in tab_row:
        # print("tabname is %s\n"%tabname)
        key_val=str(keyname.value).replace(" ","")
        # print("now is %s\n"%key_val)
        cresql=""
        if  f>0 and f<6:
            cresql="""alter table %s add %s varchar(30)"""%(tabname,key_val)
            f=f+1      
        elif f>=6:
            cresql="""alter table %s add %s text"""%(tabname,key_val)
        else :      
            cresql="""CREATE TABLE %s(
            %s CHAR(10) PRIMARY KEY)"""%(tabname,key_val)
            f=f+1
            print("create the cols...") 
        # print("%s\n"%cresql)    
        try:
            cur.execute(cresql)
            print("...")
        except:
            con.rollback()
            print("wrong!plesea check the code\n")
    if f==6:
        print("successfully finished!\n") 
    
#增加函数
def intab(tab_row):
    cur.execute("show COLUMNS from %s"%tabname)
    print("\n")
    cols= ",".join(str(i[0]).replace(" ","")for i in cur.fetchall())
    print("cols=%s"%cols)
    print("\n")
    row_value=""
    # row_value=",".join(str(i.value).replace(" ","")for i in tab_row)  #考虑到表格有Null值，无法用较为效率的循环式，改写
    for i in tab_row:  
        if i.value==None:
            row_value=row_value+", "
        else :
            row_value=row_value+"'%s'"%str(i.value)+","
    row_value=row_value.rstrip(",")    #去掉末尾多余的逗号
    print("row_value=%s"%row_value)
    print("\n")
    insql= """INSERT INTO %s(%s)
         VALUES (%s)"""%(tabname,cols,row_value)
    # print("%s\n"%insql)
    try:     
        cur.execute(insql)
        con.commit()
        return 1
    except:
        print("wrong!plesea check the code\n")
        return 0

#host,usr,password,database 为以后改写提供参数接口       
host="localhost"
usr="root"
password="006659"
database="Brca"
con=db.connect(host,usr,password,database)
cur=con.cursor()
wb=openpyxl.load_workbook('/home/tan/sju3734/project1/www/html/lib/brca.xlsx')#以后考虑改写时自行查找当前路径的Excel表。
tabname=""  #该变量为全局变量，在cretab函数以及intab函数中均有用到，请注意在函数内部中不要出现修改该变量的执行程式
f=0 #循环下标
for sheet in wb :
    tabname=wb.sheetnames[f]
    f=f+1 #此时f加1，为取出下一张excel的数据表做准备  sheet为实际的数据薄参数
    # print(tabname)
    rows=sheet.max_row
    print("max row is %d\n"%rows)
    tab_row=(x for x in sheet)
    # 以下注释代码为原始版本，易于看清执行过程，执行效率相对低，用于修改函数
    # for x in range(1,rows):            
    #     if x!=1:
    #       if intab(next(tab_row))==1:
    #           print("now is inserting row of %d\n"%x)
    #       else :
    #           print("wrong!check the code!\n")
    #           break
    #     else:
    #       print("now is cretabing row of %d\n"%x)
    #       cretab(next(tab_row))

    # #该版本为重写后，效率相对高
    next(tab_row)
    next(tab_row)
    map(intab,tab_row)
con.close()