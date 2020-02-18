'''
@Author: Tang
@Date: 2019-12-02 10:01:42
@LastEditors  : Tang
@LastEditTime : 2020-02-13 19:31:03
@Description: 
'''
import openpyxl
import sys
import os
import re
import MySQLdb as db


def add(path):
    host = "localhost"
    usr = "lcbb3734"
    password = "3734"
    database = "lcbb3734"
    con = db.connect(host, usr, password, database,
                     use_unicode=True, charset="utf8")  # 切记检查编码以及添加编码。
    cur = con.cursor()
    file_dir = os.listdir(path)  # 遍历该目录下的所有文件，文件名以数组的形式储存在file_dir中
    for i in file_dir:
        filepath = os.path.join(path, i)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        cresql = """CREATE TABLE %s(
            %s CHAR(100));""" % (i[0:-5],i[0:-5])
        cur.execute(cresql)
        print("create %s success\n"%i[0:-5])
        for x in range(1, ws.max_row+1):
            if ws['A%d' % x].value != None:
                insql = """INSERT INTO lcbb3734.%s
                VALUES ('%s');""" % (i[0:-5], ws['A%d' % x].value.replace("'", "\\'"))
                print(insql)
                print("insert %s ws[A%d].......\n"%(i[0:-5],x))
                cur.execute(insql)
                print("insert ws[A%d] success....\n"%x)
        print("all of row %d"%ws.max_row)
    con.commit()
    con.close()

if __name__ == "__main__":
    path = sys.argv[1]
    add(path)