'''
@Author: Tang
@Date: 2019-11-26 18:03:59
@LastEditors: Tang
@LastEditTime: 2019-12-03 23:03:34
@Description:
待解决：1.未解决33种肿瘤数据表的导入问题。2.PMID为空的时候，与上一行PMID一样，title与原文按上一行填充。（绝对的）          在该条件下，假如tf为空的时候，该格与上一行一样。  只有PMID号、原文、标题，其他列无数据，丢弃。
已解决：均已解决完成，目前read表因为read为mysql独占字符的原因，暂时用read1进行的代替，还有些表中的单元格是合并的，这也会造成影响，需要手动进行格式刷后再进行自动导入。
'''
import openpyxl
import sys
import os
import re
import MySQLdb as db


def addexcel(path):
    host = "localhost"
    usr = "lcbb3734"
    password = "3734"
    database = "lcbb3734"
    con = db.connect(host, usr, password, database,
                     use_unicode=True, charset="utf8")  # 切记检查编码以及添加编码。
    cur = con.cursor()
    file_dir = os.listdir(path)  # 遍历该目录下的所有文件，文件名以数组的形式储存在file_dir中
    for i in file_dir:
        filepath = os.path.join(path, i)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        ws.delete_cols(1, 2)
        print('cre table %s.....\n'%i[0:-5])
        try:
            cresql = """CREATE TABLE %s(pmid INT(20) NOT NULL,tf CHAR(30),characteristics CHAR(50),gene CHAR(30),regulation_type CHAR(40),
            hallmark CHAR(30),original_text TEXT,title TEXT);""" % i[0:-5]  # 这里是33个表的写入程序
            cur.execute(cresql)
            print('cre table %s successfully\n'%i[0:-5])
        except:
            print('the table %s already exit\n'%i[0:-5])
            continue
        for x in range(1, ws.max_row+1):
            # 2.PMID为空的时候，与上一行PMID一样，title与原文按上一行填充。（绝对的）          在该条件下，假如tf为空的时候，该格与上一行一样。  只有PMID号、原文、标题，其他列无数据，丢弃。
            print('insert table %s row of %d.......\n'%(i[0:-5],x))
            tf = ws['B%d' % x].value
            characteristics = ws['C%d' % x].value
            gene = ws['D%d' % x].value
            regulation_type = ws['E%d' % x].value
            hallmark = ws['F%d' % x].value
            flag = 0
            if tf == None:
                flag = flag + 1
            if characteristics == None:
                characteristics = 'Unkonwn'
                flag = flag + 1
            if gene == None:
                gene = 'Unkonwn'
                flag = flag + 1
            if regulation_type == None:
                regulation_type = 'Unkonwn'
                flag = flag + 1
            else:
                regulation_type = str(ws['E%d' % x].value).replace("'", "\\'")
            if hallmark == None:
                hallmark = 'Unkonwn'
                flag = flag + 1
            else:
                hallmark = str(ws['F%d' % x].value).replace("'", "\\'")
            if flag == 5:
                continue
            if ws['A%d'%x].value == None:
                y = x - 1
                ws['A%d' % x].value = ws['A%d' % y].value #防止出现连续两个ｐｍｉｄ的空值
                ws['G%d' % x].value = ws['G%d' % y].value #防止出现连续两个ｐｍｉｄ的空值
                ws['H%d' % x].value = ws['H%d' % y].value #防止出现连续两个ｐｍｉｄ的空值
                pmid = int(ws['A%d' % x].value)
                original_text = str(ws['G%d' % x].value).replace("'", "\\'")
                title = str(ws['H%d' % x].value).replace("'", "\\'")
                if tf == None:
                    tf = ws['B%d' % y].value
            else:
                pmid = int(ws['A%d' % x].value)
                original_text = str(ws['G%d' % x].value).replace("'", "\\'")
                title = str(ws['H%d' % x].value).replace("'", "\\'")
                if tf == None:
                    tf = ws['B%d' % x-1].value
            insql = """INSERT INTO lcbb3734.%s
            VALUES (%d,'%s','%s','%s','%s','%s','%s','%s');""" % (i[0:-5], pmid,tf,characteristics,gene,regulation_type,hallmark,original_text,title)
            cur.execute(insql)
            print('insert table %s row of %d successfully\n'%(i[0:-5],x))
        wb.save(filepath)
        wb.close()
        con.commit()
    con.commit()
    con.close()


if __name__ == "__main__":
    path = sys.argv[1]
    addexcel(path)
