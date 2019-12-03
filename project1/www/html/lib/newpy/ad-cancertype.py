'''
@Author: Tang
@Date: 2019-11-26 18:03:59
@LastEditors: Tang
@LastEditTime: 2019-12-03 19:27:44
@Description: 
待解决：1.未解决33种肿瘤数据表的导入问题。2.PMID为空的时候，与上一行PMID一样，title与原文按上一行填充。（绝对的）          在该条件下，假如tf为空的时候，该格与上一行一样。  只有PMID号、原文、标题，其他列无数据，丢弃。
'''
import openpyxl
import sys
import os
import re
import MySQLdb as db


def addexcel(path):
    host = "localhost"
    usr = "lcbb3734"
    password = "3734"
    database = "lcbb3734"
    con = db.connect(host, usr, password, database,
                     use_unicode=True, charset="utf8")  # 切记检查编码以及添加编码。
    cur = con.cursor()
    file_dir = os.listdir(path)  # 遍历该目录下的所有文件，文件名以数组的形式储存在file_dir中
    for i in file_dir:
        filepath = os.path.join(path, i)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        cresql = """CREATE TABLE %s(
            BriCancer CHAR(10) NOT NULL,LongCancer CHAR(100) NOT NULL);""" % i[0:-5]
        cur.execute(cresql)
        for x in range(1, ws.max_row):
            insql = """INSERT INTO lcbb3734.%s
            VALUES ('%s','%s');""" % (i[0:-5], ws['A%d' % x].value, ws['B%d' % x].value)
            cur.execute(insql)
    con.commit()
    con.close()


if __name__ == "__main__":
    path = sys.argv[1]
    addexcel(path)
