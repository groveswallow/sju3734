'''
@Author: Tang
@Date: 2019-11-26 18:03:59
@LastEditors: Tang
@LastEditTime: 2019-11-27 14:32:59
@Description: 
'''
import openpyxl
import sys
import os
import re
import MySQLdb as db


def add(path):
    host = "localhost"
    usr = "lcbb3734"
    password = "3734"
    database = "lcbb3734"
    con = db.connect(host, usr, password, database,
                     use_unicode=True, charset="utf8")  # 切记检查编码以及添加编码。
    cur = con.cursor()
    file_dir = os.listdir(path) #遍历该目录下的所有文件，文件名以数组的形式储存在file_dir中
    for i in file_dir:
        filepath = os.path.join(path, i)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        if ws.max_column <= 2:  # 这里导入的是肿瘤种类的表
            cresql = """CREATE TABLE %s(
                BriCancer CHAR(10) NOT NULL,LongCancer CHAR(100) NOT NULL);""" % i[0:-5]
            cur.execute(cresql)
            for x in range(1, ws.max_row):
                insql = """INSERT INTO lcbb3734.%s
                VALUES ('%s','%s');""" % (i[0:-5], ws['A%d' % x].value, ws['B%d' % x].value)
                cur.execute(insql)
        else:
            pass  # 这里是33个表的写入程序
    con.commit()
    con.close()


if __name__ == "__main__":
    path = sys.argv[1]
    add(path)